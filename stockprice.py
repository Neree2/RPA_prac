from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time as t
import pyautogui as pg
from openpyxl import load_workbook

loadfile='stockprice.xlsx'
myworkbook = load_workbook(loadfile)
mysheet=myworkbook.active
lastrow= len(mysheet['A'])+1#จำนวนแถวที่มีข้อมูลมีกี่แถว

stocklist = []
pricelist = []

for i in range(2, lastrow):
    stocklist.append(mysheet.cell(row=i, column=1).value)

path=r'C:\Users\ASUS\OneDrive\เดสก์ท็อป\Rpa_thing\chromedriver.exe'
service=Service(path)

option=webdriver.ChromeOptions()
option.add_argument('--start-maximuzed')

driver = webdriver.Chrome(service=service,options=option)

for s in stocklist:
    driver.get("https://www.google.com")

    stock=driver.find_element(By.NAME,'q')
    stock.send_keys(f'ราคาหุ้น {s}')
    stock.send_keys(Keys.RETURN)
    t.sleep(5)

    xpath='//*[@id="knowledge-finance-wholepage__entity-summary"]/div[3]/g-card-section/div/g-card-section/div[2]/div[1]/span[1]/span/span[1]'
    price=driver.find_element(By.XPATH, xpath).text
    pricelist.append(float(price))

    for r, p in enumerate(pricelist, 2):
        writedata = mysheet.cell(row=r, column=2)
        writedata.value=p

    img_search=pg.screenshot(region=(179,531,1000,440))#ครอปภาพ
    img_search.save('หุ้นเด้ออ.png')

    t.sleep(5)
    print(s)

driver.quit()
myworkbook.save()